#!/usr/bin/env python3
"""
IMU Data Post-processing and Visualization
Author: CZ
Date: 2025-01-09
Description: Generate beautiful heatmap and consolidated ranking table for IMU data
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Set style for beautiful academic plots
plt.style.use('seaborn-v0_8-whitegrid')
sns.set_palette("husl")


def create_beautiful_heatmap(input_file: str = 'results/imu_correlation_matrix.csv', 
                            output_file: str = 'results/imu_correlation_heatmap.png'):
    """
    Create a beautiful academic-style heatmap from IMU correlation matrix
    """
    print("="*80)
    print("Creating Beautiful IMU Correlation Heatmap")
    print("="*80)
    
    # Read correlation matrix
    corr_matrix = pd.read_csv(input_file, index_col=0)
    print(f"Loaded correlation matrix: {corr_matrix.shape}")
    
    # Convert to numeric and take absolute values
    corr_matrix = corr_matrix.astype(float).abs()
    
    # Clean column names for display (remove the long prefix)
    clean_columns = []
    for col in corr_matrix.columns:
        # Remove "GENERIC_IMU GENERIC_IMU_DATA_TLM " prefix
        clean_col = col.replace('GENERIC_IMU GENERIC_IMU_DATA_TLM ', '')
        # Further simplify if needed
        clean_col = clean_col.replace('GENERIC_IMU ', '')
        clean_columns.append(clean_col)
    
    corr_matrix.columns = clean_columns
    
    # Calculate figure size based on matrix dimensions
    n_factors = len(corr_matrix.index)
    n_stats = len(corr_matrix.columns)
    
    # Dynamic sizing with minimum and maximum constraints
    fig_width = max(12, min(24, n_stats * 1.2))
    fig_height = max(8, min(16, n_factors * 0.35))
    
    # Create figure with high DPI for academic quality
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=150)
    
    # Create custom colormap - academic style
    # Using a diverging colormap from white (0) to dark red (1)
    from matplotlib.colors import LinearSegmentedColormap
    colors = ['#ffffff', '#fff5f0', '#fee0d2', '#fcbba1', '#fc9272', 
              '#fb6a4a', '#ef3b2c', '#cb181d', '#a50f15', '#67000d']
    n_bins = 100
    cmap = LinearSegmentedColormap.from_list('custom_reds', colors, N=n_bins)
    
    # Create heatmap with enhanced aesthetics
    heatmap = sns.heatmap(corr_matrix,
                         annot=True,
                         fmt='.3f',
                         cmap=cmap,
                         vmin=0, 
                         vmax=1,
                         center=0.5,
                         square=False,
                         linewidths=0.5,
                         linecolor='#e0e0e0',
                         cbar_kws={'label': 'Absolute Correlation Coefficient |r|',
                                  'shrink': 0.8,
                                  'aspect': 20,
                                  'pad': 0.02},
                         annot_kws={'size': 8, 'weight': 'normal'})
    
    # Customize colorbar
    cbar = heatmap.collections[0].colorbar
    cbar.ax.tick_params(labelsize=10)
    cbar.ax.set_ylabel('Absolute Correlation Coefficient |r|', fontsize=12, weight='bold')
    
    # Set title with academic formatting
    ax.set_title('IMU Sensor Data Correlation Analysis\n'
                'Angular Rate & Linear Acceleration Components',
                fontsize=16, 
                weight='bold', 
                pad=20,
                fontfamily='serif')
    
    # Customize axes labels
    ax.set_xlabel('IMU Measurement Statistics', fontsize=14, weight='bold', fontfamily='serif')
    ax.set_ylabel('System Parameters / Factors', fontsize=14, weight='bold', fontfamily='serif')
    
    # Rotate labels for better readability
    plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor', fontsize=10)
    plt.setp(ax.get_yticklabels(), rotation=0, fontsize=10)
    
    # Add grid for better readability (subtle)
    ax.grid(True, which='major', alpha=0.1, linestyle='-', linewidth=0.5)
    
    # Add correlation strength indicators
    # ax.text(1.02, 0.5, 'Strong\n(|r| > 0.7)', transform=ax.transAxes,
    #         fontsize=9, va='center', ha='left', weight='bold', color='#67000d',
    #         bbox=dict(boxstyle="round,pad=0.3", facecolor='#fee0d2', alpha=0.7))
    
    # ax.text(1.02, 0.3, 'Moderate\n(0.4 < |r| < 0.7)', transform=ax.transAxes,
    #         fontsize=9, va='center', ha='left', weight='bold', color='#cb181d',
    #         bbox=dict(boxstyle="round,pad=0.3", facecolor='#fee0d2', alpha=0.5))
    
    # ax.text(1.02, 0.1, 'Weak\n(|r| < 0.4)', transform=ax.transAxes,
    #         fontsize=9, va='center', ha='left', weight='bold', color='#fb6a4a',
    #         bbox=dict(boxstyle="round,pad=0.3", facecolor='#fff5f0', alpha=0.7))
    
    # Fine-tune layout
    plt.tight_layout(rect=[0, 0, 0.95, 1])
    
    # Save with high quality
    plt.savefig(output_file, dpi=300, bbox_inches='tight', 
                facecolor='white', edgecolor='none')
    plt.close()
    
    print(f" Beautiful heatmap saved to: {output_file}")
    print(f"  Dimensions: {n_factors} factors × {n_stats} statistics")
    
    # Print correlation statistics
    flat_corr = corr_matrix.values.flatten()
    flat_corr = flat_corr[~np.isnan(flat_corr)]
    print(f"\nCorrelation Statistics:")
    print(f"  Maximum |r|: {flat_corr.max():.4f}")
    print(f"  Mean |r|: {flat_corr.mean():.4f}")
    print(f"  Median |r|: {np.median(flat_corr):.4f}")
    print(f"  Strong correlations (|r| > 0.7): {(flat_corr > 0.7).sum()}")
    print(f"  Moderate correlations (0.4 < |r| < 0.7): {((flat_corr > 0.4) & (flat_corr <= 0.7)).sum()}")
    

def create_consolidated_ranking_table(input_file: str = 'results/imu_top10_factors_analysis.csv',
                                     output_file: str = 'results/imu_consolidated_rankings.csv'):
    """
    Create a consolidated ranking table combining mean and std rankings
    """
    print("\n" + "="*80)
    print("Creating Consolidated IMU Factor Rankings")
    print("="*80)
    
    # Read the top factors analysis
    df = pd.read_csv(input_file)
    print(f"Loaded {len(df)} factor rankings")
    
    # Extract clean column names (remove prefix and suffix)
    def clean_column_name(col_name):
        # Remove the GENERIC_IMU prefix
        clean = col_name.replace('GENERIC_IMU GENERIC_IMU_DATA_TLM ', '')
        clean = clean.replace('GENERIC_IMU ', '')
        # Remove _mean or _std suffix
        clean = clean.replace('_mean', '').replace('_std', '')
        return clean
    
    # Process the data
    ranking_dict = {}
    
    for _, row in df.iterrows():
        imu_stat = row['imu_statistic']
        factor = row['factor']
        rank = row['rank']
        abs_corr = row['abs_correlation']
        
        # Clean the IMU statistic name
        base_name = clean_column_name(imu_stat)
        
        # Determine if it's mean or std
        is_mean = '_mean' in imu_stat
        is_std = '_std' in imu_stat
        
        # Initialize nested structure
        if base_name not in ranking_dict:
            ranking_dict[base_name] = {}
        if rank not in ranking_dict[base_name]:
            ranking_dict[base_name][rank] = {}
        if factor not in ranking_dict[base_name][rank]:
            ranking_dict[base_name][rank][factor] = {'mean': None, 'std': None}
        
        # Store correlation value
        if is_mean:
            ranking_dict[base_name][rank][factor]['mean'] = abs_corr
        elif is_std:
            ranking_dict[base_name][rank][factor]['std'] = abs_corr
    
    # Create consolidated table
    consolidated_data = []
    
    for imu_var in sorted(ranking_dict.keys()):
        row_data = {'IMU_Variable': imu_var}
        
        for rank in range(1, 11):  # Ranks 1-10
            if rank in ranking_dict[imu_var]:
                # Get all factors at this rank
                factors_at_rank = []
                
                for factor, corrs in ranking_dict[imu_var][rank].items():
                    mean_corr = corrs['mean']
                    std_corr = corrs['std']
                    
                    # Format factor string
                    if mean_corr is not None and std_corr is not None:
                        # Both mean and std have same rank
                        factor_str = f"{factor}({mean_corr:.4f},{std_corr:.4f})"
                    elif mean_corr is not None:
                        # Only mean
                        factor_str = f"{factor}(m:{mean_corr:.4f})"
                    else:
                        # Only std
                        factor_str = f"{factor}(s:{std_corr:.4f})"
                    
                    factors_at_rank.append(factor_str)
                
                # Join multiple factors if they exist at same rank
                row_data[f'Rank_{rank}'] = ' | '.join(factors_at_rank)
            else:
                row_data[f'Rank_{rank}'] = ''
        
        consolidated_data.append(row_data)
    
    # Create DataFrame
    result_df = pd.DataFrame(consolidated_data)
    
    # Reorder columns
    cols = ['IMU_Variable'] + [f'Rank_{i}' for i in range(1, 11)]
    result_df = result_df[cols]
    
    # Save to CSV
    result_df.to_csv(output_file, index=False)
    print(f" Consolidated rankings saved to: {output_file}")
    
    # Also create a prettier Excel version with formatting
    excel_file = output_file.replace('.csv', '.xlsx')
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='IMU Rankings', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['IMU Rankings']
        
        # Add formatting
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(50, max(12, max_length + 2))
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                      min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Not header
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    print(f" Excel version saved to: {excel_file}")
    
    # Print summary
    print(f"\nSummary:")
    print(f"  IMU variables analyzed: {len(result_df)}")
    print(f"  Variables: {', '.join(result_df['IMU_Variable'].tolist())}")
    
    # Display the table
    print("\nConsolidated Rankings Table:")
    print(result_df.to_string(index=False, max_colwidth=50))
    
    return result_df


def main():
    """Main function to run both processing steps"""
    print("IMU Data Post-Processing Tool")
    print("="*80)
    
    # Set paths
    results_dir = Path('results')
    
    # Check if results directory exists
    if not results_dir.exists():
        print(f"Error: Results directory '{results_dir}' not found!")
        print("Please run this script from the DataAnalysis directory.")
        return 1
    
    # Process correlation matrix to create beautiful heatmap
    matrix_file = results_dir / 'imu_correlation_matrix.csv'
    heatmap_file = results_dir / 'imu_correlation_heatmap.png'
    
    if matrix_file.exists():
        create_beautiful_heatmap(str(matrix_file), str(heatmap_file))
    else:
        print(f"Warning: {matrix_file} not found. Skipping heatmap generation.")
    
    # Process top factors to create consolidated ranking table
    factors_file = results_dir / 'imu_top10_factors_analysis.csv'
    rankings_file = results_dir / 'imu_consolidated_rankings.csv'
    
    if factors_file.exists():
        create_consolidated_ranking_table(str(factors_file), str(rankings_file))
    else:
        print(f"Warning: {factors_file} not found. Skipping ranking consolidation.")
    
    print("\n" + "="*80)
    print("Post-processing completed successfully!")
    print("="*80)
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())